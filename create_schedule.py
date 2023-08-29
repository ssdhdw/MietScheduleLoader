import openpyxl
from openpyxl.styles import PatternFill
import sqlite3

MODE = "group" # available modes: "group" "teacher" "room" "subject"

def merge_cells(ws):
    for column in range(3, 9):
        column_chr = chr(64 + column)
        for row in range(0, 8):
            row = row * 4 + 2
            if ws.cell(column=column, row=row).value == ws.cell(column=column, row=row + 1).value and \
                    ws.cell(column=column, row=row + 1).value == ws.cell(column=column, row=row + 2).value and \
                    ws.cell(column=column, row=row + 2).value == ws.cell(column=column, row=row + 3).value:
                text = ws.cell(column=column, row=row).value
                ws.merge_cells(f'{column_chr}{row}:{column_chr}{row + 3}')
                ws.cell(column=column, row=row, value=text)
            else:
                if ws.cell(column=column, row=row).value == ws.cell(column=column, row=row + 1).value:
                    text = ws.cell(column=column, row=row).value
                    ws.merge_cells(f'{column_chr}{row}:{column_chr}{row + 1}')
                    ws.cell(column=column, row=row, value=text).fill = PatternFill('solid', fgColor="E0E0E0")
                else:
                    ws.cell(column=column, row=row).fill = PatternFill('solid', fgColor="E0E0E0")
                    ws.cell(column=column, row=row + 1).fill = PatternFill('solid', fgColor="E0E0E0")
                if ws.cell(column=column, row=row + 2).value == ws.cell(column=column, row=row + 3).value:
                    text = ws.cell(column=column, row=row + 2).value
                    ws.merge_cells(f'{column_chr}{row + 2}:{column_chr}{row + 3}')
                    ws.cell(column=column, row=row + 2, value=text)

def table_from_group(rows):
    for row in rows:
        column = row[1] + 2
        week = row[2]
        time = row[3] - 1
        if week == 1:
            week = 2
        elif week == 2:
            week = 1
        _row = (time* 4 + week + 2)
        ws.cell(column=column,
                row=_row,
                value=f'{row[4]} | {row[5]}')
        wb.save("no_merge.xlsx")
        merge_cells(ws)
        wb.save('shedule.xlsx')

def table_from_teacher(rows):
    for row in rows:
        column = row[1] + 2
        week = row[2]
        time = row[3] - 1
        if week == 1:
            week = 2
        elif week == 2:
            week = 1
        _row = (time* 4 + week + 2)
        value = ws.cell(column=column, row=_row).value
        if value and value != "":
            value = f"{row[0]} {value}"
            ws.cell(column=column,
                    row=_row,
                    value=value)
        else:
            ws.cell(column=column,
                    row=_row,
                    value=f'{row[0]} | {row[4]} | {row[5]}')
            
def table_from_room(rows):
    for row in rows:
        column = row[1] + 2
        week = row[2]
        time = row[3] - 1
        if week == 1:
            week = 2
        elif week == 2:
            week = 1
        _row = (time* 4 + week + 2)
        value = ws.cell(column=column, row=_row).value
        if value and value != "":
            value = f"{row[0]} {value}"
            ws.cell(column=column,
                    row=_row,
                    value=value)
        else:
            ws.cell(column=column,
                    row=_row,
                    value=f'{row[0]} | {row[5]}')

# GROUP = 'УТС-12' # name
TEACHER = "Преподаватель кафедры ФизВоспитания " # teacher
# ROOM = "1202 (м)"
SUBJECT = "УТС-22"

conn = sqlite3.connect("data.db")
cur = conn.cursor()
cur.execute(f"SELECT * FROM groups WHERE name = '{SUBJECT}'")
rows = cur.fetchall()

wb = openpyxl.load_workbook("./sample.xlsx")
ws = wb.active

def create_table(rows, ws, mode):
    for row in rows:
        column = row[1] + 2
        week = row[2]
        time = row[3] - 1
        if week == 1:
            week = 2
        elif week == 2:
            week = 1
        _row = (time* 4 + week + 2)
        if mode == "group":
            ws.cell(column=column,
                    row=_row,
                    value=f'{row[4]} | {row[5]}')
        else:
            value = ws.cell(column=column, row=_row).value
            if value and value != "":
                value = f"{row[0]} {value}"
                ws.cell(column=column,
                        row=_row,
                        value=value)
            else:
                if mode == "teacher":
                    ws.cell(column=column,
                            row=_row,
                            value=f'{row[0]} | {row[4]} | {row[5]}')
                elif mode == "room":
                    ws.cell(column=column,
                            row=_row,
                            value=f'{row[0]} | {row[5]}')
                elif mode == "subject":
                        ws.cell(column=column,
                            row=_row,
                            value=f'{row[0]} | {row[4]}')
                        
create_table(rows, ws, MODE)

wb.save("no_merge.xlsx")
merge_cells(ws)
wb.save('shedule.xlsx')