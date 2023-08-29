from get_info import GetInfo
from openpyxl.styles import PatternFill
import openpyxl

def merge_cells(ws):
    for column in range(3, 9):
        column_chr = chr(64 + column)
        for row in range(0, 8):
            row = row * 4 + 2
            if ws.cell(column=column, row=row).value == ws.cell(column=column, row=row + 1).value and \
                    ws.cell(column=column, row=row + 1).value == ws.cell(column=column, row=row + 2).value and \
                    ws.cell(column=column, row=row + 2).value == ws.cell(column=column, row=row + 3).value:
                text = ws.cell(column=column, row=row).value
                ws.merge_cells(f'{column_chr}{row}:{column_chr}{row + 3}')
                ws.cell(column=column, row=row, value=text)
            else:
                if ws.cell(column=column, row=row).value == ws.cell(column=column, row=row + 1).value:
                    text = ws.cell(column=column, row=row).value
                    ws.merge_cells(f'{column_chr}{row}:{column_chr}{row + 1}')
                    ws.cell(column=column, row=row, value=text).fill = PatternFill('solid', fgColor="E0E0E0")
                else:
                    ws.cell(column=column, row=row).fill = PatternFill('solid', fgColor="E0E0E0")
                    ws.cell(column=column, row=row + 1).fill = PatternFill('solid', fgColor="E0E0E0")
                if ws.cell(column=column, row=row + 2).value == ws.cell(column=column, row=row + 3).value:
                    text = ws.cell(column=column, row=row + 2).value
                    ws.merge_cells(f'{column_chr}{row + 2}:{column_chr}{row + 3}')
                    ws.cell(column=column, row=row + 2, value=text)


def main():
    get_info = GetInfo()
    groups = get_info.get_groups()
    group_name = input("Введите название группы: ")
    if not (group_name in groups):
        print("Группа не найдена")
        return
    schedule = get_info.get_schedule(group_name)
    wb = openpyxl.load_workbook("sample.xlsx")
    ws = wb.active
    ws.title = group_name
    for i in schedule["Data"]:
        column = i["Day"] + 2

        if i["DayNumber"] == 1:
            i["DayNumber"] = 2
        elif i["DayNumber"] == 2:
            i["DayNumber"] = 1
        row = (i["Time"]["Code"] - 1) * 4 + 2 + i["DayNumber"]
        ws.cell(column=column,
                row=row,
                value=f'{i["Room"]["Name"]} | {i["Class"]["Name"]}')
    wb.save('without_merge.xlsx')
    merge_cells(ws)
    wb.save('merged.xlsx')


if __name__ == '__main__':
    main()
